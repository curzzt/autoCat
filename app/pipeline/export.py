from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional

from ..models import (
    Analysis,
    ClipSuggestion,
    JobResult,
    TranscriptSegment,
    VideoMetadata,
)
from .utils import format_clock


def export_markdown(result: JobResult, job_dir: Path) -> Path:
    md = _render_markdown(
        result.metadata, result.transcript, result.analysis, result.clips, result.warnings
    )
    path = job_dir / "report.md"
    path.write_text(md, encoding="utf-8")
    return path


def export_xlsx(result: JobResult, job_dir: Path) -> Optional[Path]:
    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError:
        return None

    wb = Workbook()

    ws_info = wb.active
    ws_info.title = "基础信息"
    md = result.metadata
    info_rows = [
        ("字段", "值"),
        ("标题", md.title),
        ("作者", md.author),
        ("时长(秒)", md.duration),
        ("点赞", md.likes),
        ("评论", md.comments),
        ("分享", md.shares),
        ("收藏", md.favorites),
        ("链接", md.url),
    ]
    for row in info_rows:
        ws_info.append(_safe_row(row))

    ws_tr = wb.create_sheet("逐字稿")
    ws_tr.append(["开始", "结束", "文本"])
    for seg in result.transcript:
        ws_tr.append([format_clock(seg.start), format_clock(seg.end), seg.text])

    ws_clip = wb.create_sheet("自动拆片表")
    ws_clip.append(
        [
            "序号",
            "开始时间",
            "结束时间",
            "切片标题",
            "爆点类型",
            "推荐理由",
            "开头字幕",
            "封面文案",
            "剪辑建议",
            "适合平台",
            "预估传播潜力",
        ]
    )
    for clip in result.clips:
        ws_clip.append(
            [
                clip.index,
                format_clock(clip.start),
                format_clock(clip.end),
                clip.title,
                clip.hotspot_type,
                clip.reason,
                clip.opening_subtitle,
                clip.cover_text,
                clip.edit_suggestion,
                "、".join(clip.platforms),
                clip.potential,
            ]
        )

    ws_copy = wb.create_sheet("标题文案")
    ws_copy.append(["序号", "切片标题", "封面文案", "开头字幕"])
    for clip in result.clips:
        ws_copy.append([clip.index, clip.title, clip.cover_text, clip.opening_subtitle])

    ws_risk = wb.create_sheet("风险提示")
    ws_risk.append(["风险点"])
    for risk in result.analysis.risks:
        ws_risk.append([risk])
    for w in result.warnings:
        ws_risk.append([w])

    path = job_dir / "clips.xlsx"
    wb.save(str(path))
    return path


def collect_exports(
    job_dir: Path, job_id: str, has_srt: bool, has_xlsx: bool
) -> Dict[str, str]:
    exports: Dict[str, str] = {
        "markdown": f"/downloads/{job_id}/report.md",
    }
    if has_srt and (job_dir / "subtitle.srt").exists():
        exports["srt"] = f"/downloads/{job_id}/subtitle.srt"
    if has_xlsx and (job_dir / "clips.xlsx").exists():
        exports["xlsx"] = f"/downloads/{job_id}/clips.xlsx"
    return exports


def _render_markdown(
    metadata: VideoMetadata,
    transcript: List[TranscriptSegment],
    analysis: Analysis,
    clips: List[ClipSuggestion],
    warnings: List[str],
) -> str:
    lines: List[str] = []
    lines.append(f"# 拆片报告：{metadata.title or '未命名视频'}")
    lines.append("")

    lines.append("## 1. 基础信息")
    lines.append("")
    lines.append(f"- 标题：{_v(metadata.title)}")
    lines.append(f"- 作者：{_v(metadata.author)}")
    lines.append(f"- 时长：{_dur(metadata.duration)}")
    lines.append(f"- 点赞：{_v(metadata.likes)}")
    lines.append(f"- 评论：{_v(metadata.comments)}")
    lines.append(f"- 分享：{_v(metadata.shares)}")
    lines.append(f"- 收藏：{_v(metadata.favorites)}")
    lines.append(f"- 链接：{_v(metadata.url)}")
    lines.append("")

    lines.append("## 2. 逐字稿")
    lines.append("")
    if transcript:
        for seg in transcript:
            lines.append(f"- `{format_clock(seg.start)}-{format_clock(seg.end)}` {seg.text}")
    else:
        lines.append("（未获取到逐字稿）")
    lines.append("")

    lines.append("## 3. 视频结构拆解")
    lines.append("")
    if analysis.structure:
        for item in analysis.structure:
            lines.append(f"- {item}")
    else:
        lines.append("（无）")
    lines.append("")

    lines.append("## 4. 爆款原因分析")
    lines.append("")
    lines.append(f"- 开头钩子：{_v(analysis.hook)}")
    lines.append(f"- 核心冲突：{_v(analysis.conflict)}")
    lines.append(f"- 情绪价值：{_v(analysis.emotion_value)}")
    lines.append(f"- 信息价值：{_v(analysis.info_value)}")
    lines.append(f"- 信任背书：{_v(analysis.trust)}")
    lines.append(f"- 转发理由：{_v(analysis.share_reason)}")
    lines.append(f"- 评论触发：{_v(analysis.comment_trigger)}")
    lines.append("")

    lines.append("## 5. 自动拆片表")
    lines.append("")
    lines.append(
        "| 序号 | 开始 | 结束 | 切片标题 | 爆点类型 | 推荐理由 | 开头字幕 | 封面文案 | 剪辑建议 | 适合平台 | 预估传播潜力 |"
    )
    lines.append("|---|---|---|---|---|---|---|---|---|---|---|")
    for clip in clips:
        lines.append(
            "| {idx} | {start} | {end} | {title} | {hotspot} | {reason} | {sub} | {cover} | {edit} | {plat} | {pot} |".format(
                idx=clip.index,
                start=format_clock(clip.start),
                end=format_clock(clip.end),
                title=_cell(clip.title),
                hotspot=_cell(clip.hotspot_type),
                reason=_cell(clip.reason),
                sub=_cell(clip.opening_subtitle),
                cover=_cell(clip.cover_text),
                edit=_cell(clip.edit_suggestion),
                plat=_cell("、".join(clip.platforms)),
                pot=_cell(clip.potential),
            )
        )
    lines.append("")

    lines.append("## 6. 标题与封面文案建议")
    lines.append("")
    for clip in clips:
        lines.append(f"- 切片 {clip.index}：标题「{_v(clip.title)}」 / 封面「{_v(clip.cover_text)}」")
    lines.append("")

    lines.append("## 7. 风险提示")
    lines.append("")
    risks = list(analysis.risks)
    if risks:
        for risk in risks:
            lines.append(f"- {risk}")
    else:
        lines.append("- 暂无特别风险提示")
    lines.append("")

    if warnings:
        lines.append("## 附：处理过程中的告警")
        lines.append("")
        for w in warnings:
            lines.append(f"- {w}")
        lines.append("")

    return "\n".join(lines)


def _v(value) -> str:
    if value is None or value == "":
        return "未获取"
    return str(value)


def _dur(value: Optional[float]) -> str:
    if value is None:
        return "未获取"
    return f"{value:.1f} 秒"


def _cell(value: str) -> str:
    return (value or "").replace("\n", " ").replace("|", "/")


def _safe_row(row):
    return [c if c is not None else "未获取" for c in row]
